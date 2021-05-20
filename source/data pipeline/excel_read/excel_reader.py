# 2021 Abdulkerim Güven
# Part of the WQPMS Project.
# Reads Excel .xlsm and .xlsx files provided by the Ministry and converts them into a format which we can use.

import multiprocessing # Gotta use them cores

import glob, pickle, os
from openpyxl import Workbook, load_workbook

from static_data import *

from datetime import datetime # datetime.fromisoformat(datetime_string: str)

import pprint # for debugging
pp = pprint.PrettyPrinter(indent = 4) # for debug

excel_files_2020 = glob.glob("./data/2020/*.xlsm") # Get a list of 2020 .xlsm files.
excel_files_2018 = glob.glob("./data/2018/Excel Files/*.xlsx") # Get a list of 2018 .xlsx files.

# Pickle utils --------------
def load_from_pickle(name):
    filename = name + ".pickle"
    data = None
    with open(filename, 'rb') as pickle_file:
        data = pickle.load(pickle_file)
    return data

def dump_to_pickle(data, name):
    filename = name + ".pickle"
    with open(filename, 'wb') as pickle_file:
        pickle.dump(data, pickle_file)
# ---------------------------

class ExcelDataReader(object):
    def __init__(self, filepath):
        self.filepath = filepath
        filename = os.path.basename(filepath)
        self.filename, file_extension = os.path.splitext(filename)
        self.workbook = load_workbook(filepath, keep_vba = True, read_only = True)
        self.data_tables = []
        self.current_worksheet      = None
        self.current_worksheet_name = None
        self.worksheet_dimensions   = None # ( Top-left (row, col), bottom-right (row, col) )
        self.current_data_table     = {}
        self.month_count = 0

        if file_extension == ".xlsm":
            self.file_year = 2020
        elif file_extension == ".xlsx":
            self.file_year = 2018
        else:
            assert False, ("unexpected file_extension:" + file_extension)

    def __del__(self):
        # This is only here because I'm curious whether Python objects get cleaned up deterministically. (in CPython)
        # As far as I can tell, objects that go out of scope are deleted immediately. Maybe RC is used?
        # print("Destructor called for", self.filename)
        self.close()

    def __str__(self):
        string = "ExcelDataReader object: Table Count {} and Filename {}".format(len(self.data_tables), self.filename)
        return string

    def close(self):
        try:
            if self.workbook is not None:
                self.workbook.close()
                self.workbook               = None
                self.current_worksheet      = None
                self.worksheet_dimensions   = None
        except:
            pass

    def check_worksheet_names(self):
        possible_sheetnames = ['Akarsu', 'Göl', 'Deniz', 'Atıksu', 'Sayfa1', 'Sayfa2', 'Sayfa3', 'RAPOR'] # "RAPOR" is only in 2020 files.
        intersection = [value for value in self.workbook.sheetnames if value in possible_sheetnames]
        return intersection

    def open_worksheet(self, ws_name):
        self.current_worksheet = self.workbook[ws_name]
        self.current_worksheet_name = ws_name

        # Dimension stuff...
        dimensions = self.current_worksheet.calculate_dimension()
        dims = dimensions.split(":")

        top_left     = dims[0]
        bottom_right = dims[1]

        top_left_row     = int(top_left[1:]) # row count
        bottom_right_row = int(bottom_right[1:]) # row count

        top_left_col     = ord(top_left[0])     - ord('A') + 1
        bottom_right_col = ord(bottom_right[0]) - ord('A') + 1

        self.current_data_table = {}
        self.worksheet_dimensions = ((top_left_col, top_left_row), (bottom_right_col, bottom_right_row))

    def get_row_names(self):
        assert (self.current_worksheet is not None), ("Use open_worksheet! Worksheet is None in " + self.filename)

        row_names = []
        top_left_column = self.worksheet_dimensions[0][0]
        row_count = self.worksheet_dimensions[1][1]

        for row_idx in range(1, row_count + 1):
            val = self.current_worksheet.cell(row = row_idx, column = top_left_column).value

            if val is not None:
                row_names.append(val)

        return row_names

    def read_row_into_table(self, row_idx, data_table):
        top_left_column = self.worksheet_dimensions[0][0]

        assert (self.current_worksheet is not None), ("Use open_worksheet! Worksheet is None in " + self.filename)
        column_name = self.current_worksheet.cell(row = row_idx, column = top_left_column).value

        if column_name in headers["numune"]: # 2020 table start
            data_temp = self.current_worksheet.cell(row = row_idx, column = top_left_column + 1).value
            data_table["Numune Kodu"] = data_temp

            if self.file_year == 2020:
                # I know too much state mutation is bad, but...
                # NOTE(ag): This should get deleted later on, or at least ignored.
                data_table["Start_Row"] = row_idx
                return 1 # Started

        if column_name in headers["bolge"]: # 2018 table start
            data_temp = self.current_worksheet.cell(row = row_idx, column = top_left_column + 1).value
            data_table["Bölge Adı"] = data_temp

            if self.file_year == 2018:
                # I know too much state mutation is bad, but...
                # NOTE(ag): This should get deleted later on, or at least ignored.
                data_table["Start_Row"] = row_idx
                return 1 # Started

        if column_name not in possible_row_names:
            data_temp = self.current_worksheet.cell(row = row_idx, column = top_left_column).value
            # print("End column_name:", column_name, "data_temp:", data_temp) # TODO(ag) This is for debugging, delete later.
            if data_temp is not None:
                data_table["Açıklama"] = data_temp
            elif (self.file_year == 2018) and (len(data_table.keys()) < 5): # 5 is just a guess, it should be enough.
                return 0 # Sometimes one row is merged with the row below, so it seems like bottom row is None. Table does not actually end in that case.
            return -1 # Ended

        if column_name in headers["yer"]:
            data_temp = self.current_worksheet.cell(row = row_idx, column = top_left_column + 1).value
            data_table["Yer"] = data_temp

        if column_name in headers["tarih"]:
            if self.file_year == 2020:
                self.month_count = 5
            else: # self.file_year == 2018:
                count = 0
                for idx in range(1, 25):
                    date_val = self.current_worksheet.cell(row = row_idx, column = top_left_column + idx).value
                    if date_val in rightmost_cells:
                        count = idx - 1
                        break

                self.month_count = count

            table_type_col = 0
            if self.file_year == 2020:
                table_type_col = 10 # It's mostly column J (10), sometimes H (8)
            else: # 2018
                table_type_col = top_left_column + self.month_count + 1

            table_type = self.current_worksheet.cell(row = data_table["Start_Row"], column = table_type_col).value

            if self.file_year == 2020 and table_type is None:
                table_type_col = 8
                table_type = self.current_worksheet.cell(row = data_table["Start_Row"], column = table_type_col).value

            assert_error_text = "Unexpected table_type: " + str(table_type) + "Row and col:" + str(data_table["Start_Row"]) + "," + str(table_type_col)
            assert (table_type in possible_table_types), (assert_error_text)

            data_table["Table_Type"] = table_type

            del data_table["Start_Row"] # We don't need it anymore.

            assert (self.month_count > 1), ("month_count must be greater than 1, weird data. -- " + self.filename)
            dates = []
            for i in range(1, self.month_count + 1):
                col_data = self.current_worksheet.cell(row = row_idx, column = top_left_column + i).value
                if col_data == "-":
                    col_data = None
                dates.append(col_data)
            data_table["Numune Alma Tarihi"] = dates

        if column_name in headers["koord"]:
            left_col = (self.current_worksheet.cell(row = row_idx, column = top_left_column + 1).value)
            right_col = (self.current_worksheet.cell(row = row_idx, column = top_left_column + 2).value)

            data_temp = None
            if right_col == None:
                if (left_col is not None) and (type(left_col) == str):
                    data_temp = left_col.replace("\n", ", ")
            else:
                data_temp = left_col + ", " + right_col

            data_table["GPS Koordinatları"] = data_temp

        if column_name in reading_types:
            assert (self.month_count > 1), "Month count is not set!"
            data = {}
            data[column_name] = []
            for i in range(1, self.month_count + 1):
                data_in_cell = self.current_worksheet.cell(row = row_idx, column = top_left_column + i).value
                if data_in_cell == "-":
                    data_in_cell = None

                data[column_name].append(data_in_cell)
            data_table[column_name] = data[column_name]

        return 0 # Neither ended nor started

    def read_tables(self, log_completion = True):
        last = -1
        table_count = 0
        row_count = self.worksheet_dimensions[1][1]

        data_table = {}
        for row_idx in range(1, row_count + 1):
            state = self.read_row_into_table(row_idx, data_table)

            if state == 1: # New table started.
                table_count += 1
            elif state == -1 and last != -1: # Table ended.
                # print("Table ended, row:", row_idx)
                # pp.pprint(data_table) # TODO(ag) This is for debugging, delete later.
                if log_completion:
                    print("--- Table {:2} done. ---".format(table_count))

                self.data_tables.append(data_table)
                data_table = {}
                self.month_count = 0

            last = state

    def generate_tsv(self) -> str:
        tsv = ""
        for table in self.data_tables:
            tsv += ExcelDataReader.__tsv_from_table(table, self.filename, table["Table_Type"])

        return tsv

    def __tsv_from_table(table, file_path, table_name) -> str:
        reading_count = len(table['Numune Alma Tarihi'])

        keys_with_single_value = []
        keys_with_multiple_values = []

        for key, value in table.items():
            if key == "Table_Type":
                continue

            if type(value) == list:
                assert (len(value) == reading_count), ("Non-uniform number of items in " + file_path + "!")
                keys_with_multiple_values.append(key)
            else:
                keys_with_single_value.append(key)

        result = ""

        for row_idx in range(reading_count):
            date_data = table['Numune Alma Tarihi'][row_idx]
            print(row_idx)
            # print("Date Data:", date_data, "Type:", type(date_data)) # debug

            if type(date_data) == str:
                date_data = datetime.fromisoformat(date_data)
            elif date_data is None:
                continue # If there is no date, there is no data. (AFAICT)

            tbl_name_date = table_name + "_" + str(date_data.year) + "_" + str(date_data.month)
            # print("tbl_name_date ->", tbl_name_date) # debug

            for key in keys_with_single_value:
                clean_key = key
                if "\n" in key:
                    clean_key = key.replace("\n", " ")

                value = str(table[key]).strip()
                line = "{}\t{}\t({})\t{}\t{}\t{}\n".format(file_path, tbl_name_date, row_idx, clean_key, "System.String", value)
                result += line
            for key in keys_with_multiple_values:
                clean_key = key
                if "\n" in key:
                    clean_key = key.replace("\n", " ")

                value = str(table[key][row_idx]).strip()
                line = "{}\t{}\t({})\t{}\t{}\t{}\n".format(file_path, tbl_name_date, row_idx, clean_key, "System.String", value)
                result += line

        return result

def get_all_data(filepath_list, log_completion = True):
    readers = []

    count = 0
    total = len(filepath_list)
    for filepath in filepath_list:
        count += 1

        edr = ExcelDataReader(filepath)

        if log_completion:
            print("Current file: {}, ({} out of {})".format(edr.filename, count, total))

        intersections = edr.check_worksheet_names()

        for ws_name in intersections:

            if log_completion:
                print("Worksheet:", ws_name)

            edr.open_worksheet(ws_name)
            edr.read_tables(log_completion)
            readers.append(edr)

        edr.close()

    return readers

# ------------- Multiprocessing stuff

def worker_edr(filepath, log_completion, worker_id):
    edr = ExcelDataReader(filepath)

    intersections = edr.check_worksheet_names()

    for ws_name in intersections:
        if log_completion:
            print("Worksheet:", ws_name)

        edr.open_worksheet(ws_name)
        edr.read_tables(log_completion)

    edr.close()
    print(worker_id, "done!")

    return edr

def get_all_data_multiprocess(filepath_list, log_completion = True):
    async_results = []

    count = 0
    total = len(filepath_list)

    params = []
    for filepath in filepath_list:
        param = (filepath, log_completion, count)
        params.append(param)
        count += 1

    edr_results = []
    with multiprocessing.Pool(processes = 16) as pool:
        edr_results = pool.starmap(worker_edr, params)

    return edr_results

# ------------- Multiprocessing stuff

def get_all_data_and_pickle(year, multiprocess = False):
    assert ((year == 2020) or (year == 2018)), "Year not valid."

    print("Year is", year)

    file_list = excel_files_2020
    pickle_name = "tables_2020"
    if year == 2018:
        file_list = excel_files_2018
        pickle_name = "tables_2018"

    readers = None
    if multiprocess:
        pickle_name = pickle_name + "_mp"
        readers = get_all_data_multiprocess(file_list)
    else:
        readers = get_all_data(file_list)

    print("Total number of files:", len(readers))
    dump_to_pickle(readers, pickle_name)
    print("Pickled the data")

    return readers

def main():
    mp = True
    # all_edrs = get_all_data_and_pickle(2018, mp) + get_all_data_and_pickle(2020, mp)
    all_edrs = load_from_pickle("tables_2018_mp") + load_from_pickle("tables_2020_mp")

    with open("stuff.tsv", "w") as fd:
        for edr in all_edrs:
            print("Writing", str(edr))
            fd.write(edr.generate_tsv() + "\n\n")

    print("Everthing done! Yay!")

if __name__ == '__main__':
    main()
